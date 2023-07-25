import pandas as pd
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.parse

wb = load_workbook('combined_result.xlsx')
ws = wb.active

company_info = []

# Base URL
base_url = "https://www.tianyancha.com/search?key="

# Start the WebDriver
driver = webdriver.Chrome()

for row in ws.iter_rows(min_row=2, max_row=1286, min_col=1, max_col=2):
    image_name = row[0].value
    name = row[1].value
    if name:
        driver.get(base_url + urllib.parse.quote(name))
        time.sleep(3)

        try:
            # Find the first search result link
            search_result = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".index_alink__zcia5.link-click")))
            company_url = search_result.get_attribute('href')

            # Get company name from website
            website_company_name = search_result.find_element(By.TAG_NAME, 'span').text

            if company_url:  # &&name==website_company_name
                company_info.append([image_name, name, company_url, website_company_name])
            else:
                company_info.append([image_name, name, '无', '无'])

        except Exception as e:
            print(f"An error occurred: {e}")
            company_info.append([image_name, name, '无', '无'])
driver.quit()

df_urls = pd.DataFrame(company_info, columns=['ImageName', 'Code', 'TYC_URL', 'CompanyName'])
df_urls.to_excel("CompanyAndTrust.xlsx", index=False)
