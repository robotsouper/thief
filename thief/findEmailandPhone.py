import pandas as pd
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wb = load_workbook('filtered_urls.xlsx')
ws = wb.active

url_email_phone = {}

driver = webdriver.Chrome()
count = 1
for row in ws.iter_rows(min_row=2, max_row=250, values_only=True):
    print("Now in row: ", count)
    count += 1
    url = row[0]
    if url != "无":
        driver.get(url)
        time.sleep(1)

    email = "无"
    phone = "无"
    try:
        phone_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '[class^="index_detail-tel__"]')))
        phone = phone_element.text

    except Exception as e:
        print(f"Cannot get phone number: {str(e)}")

    try:
        email_element = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".index_detail-email__B_1Tq")))
        email = email_element.text

    except Exception as e:
        print(f"Cannot get email: {str(e)}")

    url_email_phone[url] = [email, phone]

driver.quit()

# Create a new DataFrame with the URLs, email addresses, and phone numbers
df_contacts = pd.DataFrame.from_dict(url_email_phone, orient='index', columns=['Email', 'Phone'])

# Write the DataFrame to an Excel file
df_contacts.reset_index().rename(columns={'index': 'URL'}).to_excel("后来的我们.xlsx", index=False)
